from Cit_par import *
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import sys
import os
cwd = os.getcwd()
sys.path.insert(0, os.path.join(cwd, '../thrust'))
from get_thrust_dat import get_pressure, get_mach
sys.path.insert(0, os.path.join(cwd, '../mass_balance_report'))
from weight import *

'Initialize global variables'
alpha0theoretical = -1.2   #deg
cmalphatheretical = 5.084*np.pi/180. #rad
CD0theoretical = 0.04
etheoretical = 0.8


# Data from thrust.exe
Thrust = np.genfromtxt('../thrust/thrust1.dat') # Used for CD
Thrust2 = np.genfromtxt('../thrust/thrust2.dat') # Used for equivalent elevator deflection
Thrust2s = np.genfromtxt('../thrust/thrust2s.dat') # Used for equivalent elevator (standard thrust)
Thrust3 = np.genfromtxt('../thrust/thrust3.dat') # Used for Cmdelta Cmalpha


def get_data(measurement=1):
    Temp0 = 288.15  # temperature at sea level in ISA [K]
    lambda_1 = -0.0065  # temperature gradient in ISA [K/m]
    if measurement==1:
        ran=['28','33']
        letters=['D','E','G','H','J','I']
    elif measurement==2:
        ran=['59','65']
        letters=['D','E','J','K','M','L']

    elif measurement==3:
        ran=['75','76']
        letters=['D','E','J','K','M','L']
    wb = load_workbook("../data/Common_Post_Flight_Datasheet_Flight.xlsx", read_only=True)
    sheets = wb.sheetnames
    data_excel = wb[sheets[0]]

    hp = np.array([[float(i.value) for i in j] for j in data_excel[letters[0] + ran[0]:letters[0] + ran[1]]]) * 0.3048
    Vc = (np.array([[float(i.value) for i in j] for j in data_excel[letters[1] + ran[0]:letters[1] + ran[1]]])-2) * 0.514444
    pressure=np.array([[get_pressure(hp[i,0])] for i in range(hp.shape[0])])
    mach=np.array([[get_mach(Vc[i,0],pressure[i,0])] for i in range(hp.shape[0])])
    temp_i = np.array(
        [[float(i.value) for i in j] for j in data_excel[letters[4] + ran[0]:letters[4] + ran[1]]]) + 273.15
    temp = np.array([[(temp_i[i, 0] / (1 + 0.2 * 0.9 * mach[i, 0] ** 2))] for i in
                     range(hp.shape[0])])
    a2 = temp*1.4*287.05
    fuel_used = np.array([[float(i.value) for i in j] for j in data_excel[letters[5] + ran[0]:letters[5] + ran[1]]]) / 2.20462
    Fe = np.array([[float(i.value) for i in j] for j in data_excel['I' + ran[0]:'I' + ran[1]]])
    delta_e = np.array([[float(i.value) for i in j] for j in data_excel['G' + ran[0]:'G' + ran[1]]])*pi/180
    delta_tr = np.array([[float(i.value) for i in j] for j in data_excel['H' + ran[0]:'H' + ran[1]]])*pi/180
    aoa_0= np.array([[float(i.value) for i in j] for j in data_excel['F' + ran[0]:'F' + ran[1]]])*pi/180

    W=(initial_mass-fuel_used)*g
    V = np.multiply(mach,np.sqrt(a2))
    rho = np.divide(pressure,temp)/287.05
    return temp, rho, W, V, hp, Fe, delta_e, delta_tr, aoa_0


def get_fit(x, y, degree, scale):
    fitting = np.polyfit(x, y, degree)
    point = 0
    for i in range(int(degree)+1):
        point += fitting[i]*scale**(degree-i)
    return point


def get_time(hour, minutes, sec):
    time = hour * 60. * 60. + minutes * 60. + sec
    return time


def get_CL(measurement=1):
    temp, rho, W, V, hp, Fe, delta_e, delta_tr, aoa_0 = get_data(measurement=measurement)
    CL=2*np.divide(W, np.multiply(rho, np.square(V)))/S
    return list(CL.reshape((1, hp.shape[0])))[0]


def get_CL_alpha_slope(include_theoretical_trend=False):
    temp, rho, W, V, hp, Fe, delta_e, delta_tr, aoa_0=get_data(measurement=1)
    alpha_CL = np.zeros((aoa_0.shape[0], 2))
    CL=get_CL()
    for i in range(aoa_0.shape[0]):
        alpha_CL[i][0] = aoa_0[i,0]
        alpha_CL[i][1] = CL[i]
    slope = np.polyfit(alpha_CL[:, 0], alpha_CL[:, 1], 1)[0] # before transforming in degrees
    angles = np.degrees(alpha_CL[:,0])
    scale = np.linspace(-2, 12, 120) #np.degrees(alpha_CL[:,0])
    CL_theoretical = [cmalphatheretical*(i-alpha0theoretical) for i in scale]
    CL_fit = [get_fit(angles, alpha_CL[:,1], 1, scale[i]) for i in range(np.shape(scale)[0])]
    #plt.figure(dpi=100)
    plt.plot(scale, CL_fit, color= '#1c61a5', label= 'Experimental trend')
    if include_theoretical_trend:
        plt.plot(scale, CL_theoretical, linestyle='dashed', color='green', label= 'Theoretical trend')
    plt.plot(np.degrees(alpha_CL[:, 0]), alpha_CL[:, 1], 'o', color='#fc6910', label= 'Measurement points')
    plt.minorticks_on()
    plt.grid(which='major', linestyle='-', color='gray')
    plt.grid(which='minor', linestyle='dotted', color='lightgray')
    plt.xlabel('$\\alpha$ [deg]')
    plt.ylabel('$C_L$ [-]')
    plt.title('$C_L - \\alpha$')
    plt.legend()
    plt.savefig(f'../plots/CL_alpha_include_theoretical_{str(include_theoretical_trend).lower()}.png', dpi=600)
    plt.show()
    return slope


def get_verification(positions, funct, var):

    machs = np.zeros((np.shape(positions)[0], 2))
    for i in time_measurements[:48220]:
        mach_calc = funct(i)
        idx = get_time2idx(i)
        machs[idx][0] = mach_calc
        machs[idx][1] = get_arbitrary_variable(i, var)

    plt.plot(positions, machs[:,0], label='analytical')
    plt.plot(positions, machs[:, 1], label='experimental')
    # plt.title('Mach verification plot')
    plt.legend()
    plt.show()


def get_thrust(positions):
    for i in range(np.shape(positions)[0]):
        a = get_arbitrary_variable(positions[i], hp0)
        b = get_arbitrary_variable(positions[i], mach)
        T_stat = get_arbitrary_variable(positions[i], Tstatic)
        T_ISA = Temp0 + lambda_1 * a
        c = T_stat - T_ISA
        d = get_arbitrary_variable(positions[i], rh_engine_FMF) #flow_l_r[i, 1]
        e = get_arbitrary_variable(positions[i], lh_engine_FMF) #flow_l_r[i, 0]
        print("%.4f %.4f %.4f %.4f %.4f" % (a, b, c, d, e))


def get_CD():
    temp, rho, W, V, hp, Fe, delta_e, delta_tr, aoa_0 = get_data(measurement=1)
    tot_thrust = np.sum(Thrust, axis=1)
    qs = np.multiply(rho, np.square(V)).transpose()*S
    CD_array = 2 * np.divide(tot_thrust, qs)
    CD_array = list(CD_array)[0]
    return CD_array


def get_CD_CL_2_CD0_e(plot=True, include_theoretical_trend=False):
    CD_array = get_CD()

    # CL2_array = np.square([get_CL(positions[i]) for i in range(np.shape(positions)[0])])
    CL2_array = np.square(get_CL(measurement=1))
    e_coeff, CD0 = np.polyfit(CL2_array, CD_array, 1)
    e = 1/(e_coeff*Ar*pi)
    scale = np.linspace(0.1, 1, 120)
    CDtheoretical = [CD0theoretical + CL2_array[i]/(np.pi*Ar*etheoretical) for i in range(np.shape(CL2_array)[0])]
    CD_fit = [get_fit(CL2_array, CD_array, 1, scale[i]) for i in range(np.shape(scale)[0])]
    if plot:
        plt.plot(CD_fit, scale, color='#1c61a5', label='Experimental trend')
        plt.plot(CD_array, CL2_array, 'o', color='#fc6910', label='Measurement points')
        if include_theoretical_trend:
            plt.plot(CDtheoretical, CL2_array, linestyle='dashed', color='green', label='Theoretical trend')
        plt.minorticks_on()
        plt.grid(which='major', linestyle='-', color='gray')
        plt.grid(which='minor', linestyle='dotted', color='lightgray')
        plt.legend()
        plt.ylabel('$C_L^2$ [-]')
        plt.xlabel('$C_D$ [-]')
        plt.title('$C_L^2 - C_D$')
        plt.grid(True)
        plt.savefig(f'../plots/CL2_CD_include_theoretical_{str(include_theoretical_trend).lower()}.png', dpi=600)
        plt.show()
    return CL2_array, e, CD0, CDtheoretical


def plot_CD_alpha(include_theoretical_trend=False):
    _, _, _, CDtheoretical = get_CD_CL_2_CD0_e(plot=False)
    CD = get_CD()
    temp, rho, W, V, hp, Fe, delta_e, delta_tr, aoa_0=get_data(measurement=1)

    alpha_CD = np.zeros((aoa_0.shape[0], 2))
    for i in range(aoa_0.shape[0]):
        alpha_CD[i][0] = aoa_0[i, 0]
        alpha_CD[i][1] = CD[i]
    angles = np.degrees(alpha_CD[:, 0])
    scale = np.linspace(1.5, 12, 120)
    CD_fit = [get_fit(angles, alpha_CD[:, 1], 2, scale[i] ) for i in range(np.shape(scale)[0])]
    CD_theoretical = [get_fit(angles, CDtheoretical, 2, scale[i] ) for i in range(np.shape(scale)[0])]
    plt.plot(scale, CD_fit, color='#1c61a5', label='Experimental trend')
    plt.plot(angles, alpha_CD[:, 1], 'o', color='#fc6910', label='Measurement points')
    if include_theoretical_trend:
        plt.plot(scale, CD_theoretical, linestyle='dashed', color='green', label='Theoretical trend')
    plt.minorticks_on()
    plt.grid(which='major', linestyle='-', color='gray')
    plt.grid(which='minor', linestyle='dotted', color='lightgray')
    plt.legend()
    plt.xlabel('$\\alpha$ [deg]')
    plt.ylabel('$C_D$ [-]')
    plt.title('$C_D - \\alpha$')
    plt.savefig(f'../plots/CD_alpha_include_theoretical_{str(include_theoretical_trend).lower()}.png', dpi=600)
    plt.show()


def get_CD_CL(include_theoretical_trend=False):
    CL2, e, CD0, _ = get_CD_CL_2_CD0_e(plot=False, include_theoretical_trend=False)
    CL=np.sqrt(CL2)
    scale = np.linspace(0, 1, 120)
    CD = CD0 + CL2/(pi*e*Ar)
    CD_fit = [get_fit(CL, CD, 2, scale[i]) for i in range(np.shape(scale)[0])]
    CDtheory= CD0theoretical + np.square(scale)/(pi*etheoretical*Ar)
    plt.xlim([0, 0.12])
    plt.plot(CD, CL, 'o', color='#fc6910', label='Measurement points')
    if include_theoretical_trend:
        plt.plot(CDtheory, scale, linestyle='dashed', color='green', label='Theoretical trend' )
    plt.plot(CD_fit, scale, color='#1c61a5', label='Experimental trend')
    plt.xlabel('$C_D$ [-]')
    plt.ylabel('$C_L$ [-]')
    plt.title('$C_L - C_D$')
    plt.minorticks_on()
    plt.grid(which='major', linestyle='-', color='gray')
    plt.grid(which='minor', linestyle='dotted', color='lightgray')
    plt.legend()
    plt.savefig(f'../plots/CL_CD_include_theoretical_{str(include_theoretical_trend).lower()}.png', dpi=600)
    plt.show()


def get_Reynolds(time):
    V = get_arbitrary_variable(time, V0)
    # T_isa, rho_isa, p_isa = get_ISA(time)
    T = get_arbitrary_variable(time, Tstatic)
    b = 1.458e-8
    S = 110.4
    mu = b*sqrt(T**3.)/(T+S)
    # Re = rho_isa * V * c / mu
    kinematic_viscosity = 13.53e-6  # 1.2462E-5
    Re = V * c / kinematic_viscosity
    # Re = V * c * 70000
    return Re


def get_equivalent_elevator(Ver):
    temp, rho, W, V, hp, Fe, delta_e, delta_tr, aoa_0=get_data(measurement=2)
    deltas = delta_e    # list(delta_e.transpose())[0]
    rhos = rho          # list(rho.transpose())[0]
    Vs = V              # list(V.transpose())[0]
    d = 27. * 0.0254    # meter
    # coeff = cmtc * 2. / (Cmde * d**2.)
    Ver=np.array(Ver).reshape((7,1))
    V=np.array(V).reshape((7,1))
    sum_thrust=np.sum(Thrust2, axis=1).reshape((7,1))
    dimension=np.multiply(rhos*d**2.,np.square(Ver))
    tot_thrust = 2*np.divide(sum_thrust,dimension)
    sum_thrust=np.sum(Thrust2s, axis=1).reshape((7,1))
    dimension=np.multiply(rhos*d**2.,np.square(V))
    tot_thrust_s = 2*np.divide(sum_thrust,dimension)
    equivalent_deflection = deltas.transpose() - 1/Cmde*cmtc*(tot_thrust_s - tot_thrust)
    return list(equivalent_deflection)[0]


def plot_elevator_velocity():
    temp, rho, W, V, hp, Fe, delta_e, delta_tr, aoa_0=get_data(measurement=2)
    WS = 60500
    ratio=WS/W
    Ver = sorted(list(np.multiply(V,np.sqrt(rho/rho0*ratio)).transpose())[0])
    equivalent_elevator = np.degrees(sorted(get_equivalent_elevator(Ver)))
    scale = np.linspace(min(Ver), max(Ver), 120)
    elevator_fit = [get_fit(Ver, equivalent_elevator, 2, scale[i]) for i in range(np.shape(scale)[0])]
    #plt.plot(dpi=100)
    plt.plot(scale, elevator_fit, color='#1c61a5', label= 'Experimental curve')
    plt.plot(Ver, equivalent_elevator, 'o', color='#fc6910', label='Measurement points')
    plt.gca().invert_yaxis()
    plt.minorticks_on()
    plt.grid(which='major', linestyle='-', color='gray')
    plt.grid(which='minor', linestyle='dotted', color='lightgray')
    #plt.axhline(linewidth=1, color='#000000', )
    plt.ylabel('$\\delta_{e}^*$ [deg]')
    plt.xlabel('$\\tilde{V}_e$ [m/s]')
    plt.title('$\\delta_{e}^* - \\tilde{V}_e$' )
    plt.legend()
    plt.savefig('../plots/elevator_velocity.png', dpi=600)
    plt.show()


def plot_force_velocity():
    temp, rho, W, V, hp, Fe, delta_e, delta_tr, aoa_0 = get_data(measurement=2)
    WS = 60500
    ratio = WS / W
    equivalent_force = list(np.multiply(Fe, ratio).transpose())[0]
    Ver = list(np.multiply(V,np.sqrt(rho/rho0*ratio)).transpose())[0]
    scale = np.linspace(min(Ver), max(Ver), 120)
    force_fit = [get_fit(Ver, equivalent_force, 2, scale[i]) for i in range(np.shape(scale)[0])]
    plt.plot(scale, force_fit, color= '#1c61a5', label= 'Experimental curve')
    plt.plot(Ver, equivalent_force, 'o', color='#fc6910', label= 'Measurement points')
    plt.gca().invert_yaxis()
    plt.minorticks_on()
    plt.grid(which='major', linestyle='-', color='gray')
    plt.grid(which='minor', linestyle='dotted', color='lightgray')
    #plt.axhline(linewidth=1, color='#000000', )
    plt.ylabel('$F_{e}^*$ [N]')
    plt.xlabel('$\\tilde{V}_e$ [m/s]')
    plt.title('$F_{e}^* - \\tilde{V}_e$')
    plt.legend()
    plt.savefig('../plots/force_velocity.png', dpi=600)
    plt.show()


def plot_velocity(positions2, var):
    deltas = [get_arbitrary_variable(positions2[i], var) for i in range(np.shape(positions2)[0])]
    velocity = [get_arbitrary_variable(positions2[i], V0) for i in range(np.shape(positions2)[0])]
    plt.plot(velocity, np.degrees(deltas), 'ro')
    plt.gca().invert_yaxis()
    plt.grid(which='major', linestyle=':', color='gray')
    plt.axhline(linewidth=1, color= '#000000', )

    plt.show()


def getcmde():
    temp, rho, W, V, hp, Fe, delta_e, delta_tr, aoa_0=get_data(measurement=3)
    diff_elevator = delta_e[1,0]-delta_e[0,0]
    tot_thrust = np.sum(Thrust3, axis=1)
    d = 27 * 0.0254     # meter
    T2 = tot_thrust[1]  # /(rho[0,0]*S*V[0,0]**2*0.5*d**2) #4094.4
    T1 = tot_thrust[0]  # /(rho[1,0]*S*V[1,0]**2*0.5*d**2) #4097.98
    diff_thrust = T2-T1
    CN = get_CL(measurement=3)[0]

    # Get cg from interpolation with mass

    'First measurement'
    fuel_used = initial_mass - W[0, 0] / g
    fuel_mass = 2700 / 2.20462 - fuel_used
    moment_payload = 4163.2378 #3734.085
    xcg1 = get_xcg(fuel_mass, moment_payload)
    'Second measurement'
    fuel_used = (initial_mass - W[1, 0]/g)
    fuel_mass = 2700 / 2.20462 - fuel_used
    new_arm = 131 * 2.54 / 100
    old_arm = 288 * 2.54 / 100
    w_pax = 86
    moment_payload = moment_payload + new_arm * w_pax - old_arm * w_pax
    xcg2 = get_xcg(fuel_mass, moment_payload)
    diff_xcg = xcg2 - xcg1

    coeff_cmde1 = -1 / diff_elevator
    coeff_cmde2 = CN * diff_xcg / c
    coeff_cmde3 = cmtc * diff_thrust
    cmde = coeff_cmde1 * coeff_cmde2
    return cmde


def getcmalpha(cmde):
    temp, rho, W, V, hp, Fe, delta_e, delta_tr, aoa_0=get_data(measurement=2)
    de_da=np.polyfit(list(aoa_0.transpose())[0],list(delta_e.transpose())[0],1)[0]
    temp, rho, W, V, hp, Fe, delta_e, delta_tr, aoa_0=get_data(measurement=3)
    pressure = get_pressure(hp)
    v01 = V[0, 0]
    v02 = V[1, 0]
    d = 27 * 0.0254  # meter
    T, rho1, p = temp[0, 0], rho[0, 0], pressure[0, 0]
    T, rho2, p = temp[1, 0], rho[1, 0], pressure[1, 0]
    tot_thrust = np.sum(Thrust3, axis=1)
    T1 = tot_thrust[0] / (rho1 * (v01 ** 2.)*(d ** 2.)) #4094.4
    T2 = tot_thrust[1] / (rho2 * (v02 ** 2.)*(d ** 2.)) #4097.98
    diff_thrust = T2 - T1
    coeff_cmalpha1 = cmde * de_da
    cmalpha = -coeff_cmalpha1
    return cmalpha


if __name__ == '__main__':
    # Specify if the plots should include the theoretical trend or not
    include_theoretical_trend = True

    get_CD_CL(include_theoretical_trend)
    CL2, e, CD0, _ = get_CD_CL_2_CD0_e(plot=True, include_theoretical_trend=include_theoretical_trend)
    print(f'Calculated Oswald factor: {e}')
    print(f'Calculated CD0: {CD0}')
    slope = get_CL_alpha_slope(include_theoretical_trend)
    print(f'Calculated Clalpha: {slope}')
    plot_force_velocity()
    plot_elevator_velocity()
    plot_CD_alpha(include_theoretical_trend)
    cmde = getcmde()
    cmalpha = getcmalpha(cmde)
    print(f'Calculated Cmde: {cmde}')
    print(f'Calculated Cmalpha: {cmalpha}')
