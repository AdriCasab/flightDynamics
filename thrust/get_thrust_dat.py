import os
from openpyxl import load_workbook
import numpy as np


def get_pressure(hp):
    lambda_1 = -0.0065  # temperature gradient in ISA [K/m]
    Temp0 = 288.15      # temperature at sea level in ISA [K]
    R = 287.05          # specific gas constant [m^2/sec^2K]
    g = 9.81            # [m/sec^2] (gravity constant)
    pressure0 = 101325  # air pressure at sea level [Pa]
    p = pressure0 * (1 + lambda_1 * hp / Temp0) ** (-1 * g/(lambda_1 * R))
    return p


def get_mach(Vc, p):
    air_ratio = 1.4
    rho0 = 1.225
    pressure0 = 101325
    member1 = 2. / (air_ratio - 1.)
    member2 = -1. + ((1. + ((air_ratio - 1.) * rho0 * (Vc**2.) / (2. * air_ratio * pressure0)))**(air_ratio / (air_ratio - 1.)))
    member3 = (1 + pressure0 / p * member2)**((air_ratio-1)/air_ratio) - 1
    M = np.sqrt(member1 * member3)
    return M


def get_thrust_dat(measurement=1,cond='n'):
    Temp0 = 288.15      # temperature at sea level in ISA [K]
    lambda_1 = -0.0065  # temperature gradient in ISA [K/m]
    mfs = 0.048

    if measurement==1:
        ran=['28','33']
        letters=['D','E','G','H','J']
    elif measurement == 2:
        ran=['59', '65']
        letters=['D', 'E', 'J', 'K', 'M']

    elif measurement == 3:
        ran=['75', '76']
        letters=['D', 'E', 'J', 'K', 'M']

    wb = load_workbook("../data/Common_Post_Flight_Datasheet_Flight.xlsx", read_only=True)
    sheets = wb.sheetnames
    data_excel = wb[sheets[0]]

    hp = np.array([[float(i.value) for i in j] for j in data_excel[letters[0] + ran[0]:letters[0] + ran[1]]]) * 0.3048
    Vc = (np.array([[float(i.value) for i in j] for j in data_excel[letters[1] + ran[0]:letters[1] + ran[1]]]) - 2) * 0.514444
    fl = np.array([[(cond != 's')*float(i.value) for i in j] for j in data_excel[letters[2]+ran[0]:letters[2]+ran[1]]]) * 0.000125998 + (cond =='s') * mfs
    fr = np.array([[(cond != 's')*float(i.value) for i in j] for j in data_excel[letters[3]+ran[0]:letters[3]+ran[1]]]) * 0.000125998 + (cond =='s') * mfs
    pressure=np.array([[get_pressure(hp[i, 0])] for i in range(hp.shape[0])])

    mach = np.array([[get_mach(Vc[i, 0], pressure[i, 0])] for i in range(hp.shape[0])])
    temp_i = np.array([[float(i.value) for i in j] for j in data_excel[letters[4] + ran[0]:letters[4] + ran[1]]]) + 273.15
    temp = np.array([[(temp_i[i, 0]/(1 + 0.2 * mach[i, 0] ** 2) - (Temp0 + lambda_1 * hp[i, 0]))] for i in range(hp.shape[0])])
    a = np.array([hp, mach, temp, fl, fr]).transpose()

    with open('matlab.dat', 'wb') as f:
        for line in a:
            np.savetxt(f, line, fmt='%.6f')

    cwd = os.getcwd()
    cmd1 = 'cd ' + cwd
    os.system(cmd1)
    os.system('thrust.exe')
    title = 'thrust' + str(measurement) + cond+'.dat'
    os.rename('thrust.dat', title)

    return


if __name__ == "__main__":
    # get_thrust_dat(measurement=1, cond='')
    # get_thrust_dat(measurement=2, cond='')
    # get_thrust_dat(measurement=2, cond='s')
    get_thrust_dat(measurement=3, cond='')
